from rest_framework import viewsets, permissions, filters, pagination
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Direction
from .serializers import DirectionSerializer
from django.shortcuts import render

class CustomPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

class DirectionViewSet(viewsets.ModelViewSet):
    queryset = Direction.objects.all()
    serializer_class = DirectionSerializer
    from apps.accounts.granular_permissions import GranularPermission
    permission_classes = [permissions.IsAuthenticated, GranularPermission]
    module_name = 'directions'
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'code']

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log_action('create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        changed_fields = [k for k in self.request.data.keys() if k not in ['id', 'csrfmiddlewaretoken']]
        extra = f"O'zgarganlar: {', '.join(changed_fields)}" if changed_fields else "Tahrirlandi"
        self._log_action('update', instance, extra_details=extra)

    def perform_destroy(self, instance):
        self._log_action('delete', instance)
        instance.delete()

    def _log_action(self, action_type, instance, extra_details=None):
        from apps.logs.models import SystemLog
        
        action_map = {
            'create': "Yo'nalish yaratildi",
            'update': "Yo'nalish tahrirlandi",
            'delete': "Yo'nalish o'chirildi"
        }
        
        details = f"Yo'nalish: {instance.name} (Code: {instance.code})"
        if action_type == 'delete':
             details = f"Yo'nalish: {instance.name} (Code: {instance.code})"
            
        if extra_details:
             details += f". {extra_details}"
            
        ip = self.request.META.get('REMOTE_ADDR')
        
        SystemLog.objects.create(
            user=self.request.user,
            action=action_map.get(action_type, action_type),
            details=details,
            ip_address=ip
        )

    @action(detail=False, methods=['post'])
    def import_data(self, request):
        import openpyxl
        from rest_framework.parsers import MultiPartParser
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Fayl tanlanmadi'}, status=400)
            
        try:
            wb = openpyxl.load_workbook(file_obj)
            sheet = wb.active
            
            count = 0
            errors = []
            
            # Assumptions: 
            # Row 1 is header. 
            # Columns: Nomi (Name), Kod (Code)
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]: # Skip empty name
                    continue
                    
                name = str(row[0]).strip()
                code = str(row[1]).strip() if len(row) > 1 and row[1] else name.upper()[:10]
                
                # Check duplication
                if Direction.objects.filter(code=code).exists():
                    # Update or Skip? Let's skip or update name.
                    # For now just skip to avoid overwrite complex logic
                    d = Direction.objects.get(code=code)
                    if d.name != name:
                        d.name = name
                        d.save()
                        self._log_action('update', d, 'Import orqali yangilandi')
                    continue
                    
                instance = Direction.objects.create(name=name, code=code)
                self._log_action('create', instance, 'Import orqali yaratildi')
                count += 1
                
            return Response({'status': 'success', 'count': count})
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def sample_file(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Namuna"
        
        # Headers
        ws.append(['Nomi', 'Kod'])
        
        # Sample Data
        ws.append(['Dasturiy Injiniring', 'DI-2024'])
        ws.append(['Axborot Xavfsizligi', 'AX-2024'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=yo_nalishlar_namuna.xlsx'
        
        wb.save(response)
        return response

def direction_list_view(request):
    return render(request, 'crud_list.html', {'page': 'directions'})
