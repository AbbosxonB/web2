from rest_framework import viewsets, permissions, filters, status, decorators, pagination
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from .models import TestResult
from .serializers import TestResultSerializer
from docxtpl import DocxTemplate
from django.conf import settings
from django.utils import timezone
from django.utils.encoding import escape_uri_path
import os
from apps.accounts.granular_permissions import GranularPermission

class CustomPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

class TestResultViewSet(viewsets.ModelViewSet):
    queryset = TestResult.objects.all()
    serializer_class = TestResultSerializer
    permission_classes = [permissions.IsAuthenticated, GranularPermission]
    module_name = 'results'
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'test', 'student__group']
    search_fields = ['student__full_name', 'test__title', 'student__group__name']

    def get_queryset(self):
        user = self.request.user
        if user.role == 'student':
            # Students only see their own results
            # Students only see their own results
            return TestResult.objects.filter(student__user=user).order_by('-id')
        
        queryset = TestResult.objects.all().select_related('student', 'test', 'student__group').order_by('-id')

        # Date Filtering
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(completed_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(completed_at__date__lte=end_date)
            
        # Student Filters
        course = self.request.query_params.get('course')
        edu_form = self.request.query_params.get('education_form')
        direction = self.request.query_params.get('direction')
        
        if course:
            queryset = queryset.filter(student__course=course)
        if edu_form:
            queryset = queryset.filter(student__education_form=edu_form)
        if direction:
            queryset = queryset.filter(student__direction__icontains=direction)
            
        return queryset

    def perform_destroy(self, instance):
        user = self.request.user
        if user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Natijani o'chirish uchun faqat Admin huquqi talab qilinadi.")
        self._log_action('delete', instance)
        instance.delete()

    @action(detail=True, methods=['post'])
    def allow_retake(self, request, pk=None):
        user = request.user
        if user.role not in ['admin', 'dean']:
             return Response({'error': 'Huquq yo\'q'}, status=status.HTTP_403_FORBIDDEN)
        
        result = self.get_object()
        result.can_retake = True
        result.retake_granted_by = user
        result.save()
        
        self._log_action('retake', result)
        
        return Response({'status': 'retake_granted'})

    @action(detail=False, methods=['post'])
    def bulk_action(self, request):
        user = request.user
        if user.role not in ['admin', 'dean']:
            return Response({'error': 'Huquq yo\'q'}, status=status.HTTP_403_FORBIDDEN)
        
        action = request.data.get('action')
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({'error': 'Hech narsa tanlanmadi'}, status=status.HTTP_400_BAD_REQUEST)
            
        queryset = TestResult.objects.filter(id__in=ids)
        
        if action == 'delete':
            # Check delete permission if needed, but admin/dean is already checked
            queryset.delete()
            self._log_action('bulk_action', extra_details=f"{len(ids)} ta natija o'chirildi")
            return Response({'status': 'deleted', 'count': len(ids)})
            
        elif action == 'retake':
            queryset.update(can_retake=True, retake_granted_by=user)
            self._log_action('bulk_action', extra_details=f"{len(ids)} ta natijaga qayta topshirish ruxsati berildi")
            return Response({'status': 'retake_granted', 'count': len(ids)})
            
        return Response({'error': 'Noto\'g\'ri amal'}, status=status.HTTP_400_BAD_REQUEST)

    def _log_action(self, action_type, instance=None, extra_details=None):
        from apps.logs.models import SystemLog
        
        action_map = {
            'delete': "Natija o'chirildi",
            'retake': "Qayta topshirishga ruxsat berildi",
            'bulk_action': "Ommaviy amal (Natijalar)"
        }
        
        details = ""
        if instance:
            student_name = instance.student.full_name if instance.student else "Noma'lum"
            test_title = instance.test.title if instance.test else "Noma'lum"
            details = f"Natija: {student_name} - {test_title}"
            
        if extra_details:
             details = f"{details}. {extra_details}" if details else extra_details
            
        ip = self.request.META.get('REMOTE_ADDR')
        
        SystemLog.objects.create(
            user=self.request.user,
            action=action_map.get(action_type, action_type),
            details=details,
            ip_address=ip
        )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse

        user = request.user
        if user.role not in ['admin', 'dean']:
             return Response({'error': 'Huquq yo\'q'}, status=status.HTTP_403_FORBIDDEN)

        # Apply filters
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"

        # Headers
        headers = ["ID", "Talaba", "Guruh", "Test", "Ball", "Maks. Ball", "Foiz", "Holat", "Tugagan vaqti", "Sarflangan vaqt", "Sana"]
        ws.append(headers)

        from django.utils import timezone
        
        for result in queryset:
            student_name = result.student.full_name if result.student else "Noma'lum"
            group_name = result.student.group.name if (result.student and result.student.group) else "-"
            test_title = result.test.title if result.test else "-"
            
            # Calculate duration
            duration_str = "-"
            if result.completed_at and result.started_at:
                diff = result.completed_at - result.started_at
                total_seconds = int(diff.total_seconds())
                if total_seconds < 60:
                    duration_str = f"{total_seconds} sek"
                else:
                    minutes = total_seconds // 60
                    seconds = total_seconds % 60
                    duration_str = f"{minutes} min {seconds} sek"

            # Localize times
            completed_str = "-"
            if result.completed_at:
                completed_str = timezone.localtime(result.completed_at).strftime("%d.%m.%Y %H:%M")

            started_str = "-"
            if result.started_at:
                started_str = timezone.localtime(result.started_at).strftime("%d.%m.%Y %H:%M")

            ws.append([
                result.id,
                student_name,
                group_name,
                test_title,
                result.score,
                result.max_score,
                f"{result.percentage:.1f}%",
                result.status,
                completed_str,
                duration_str,
                started_str
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=natijalar.xlsx'
        wb.save(response)
        return response


# MOVED OUTSIDE THE CLASS - This is the fix!
def export_docx_view(request):
    """Standalone function-based view for export"""
    user = request.user
    
    if not user.is_authenticated:
        return HttpResponse('Unauthorized', status=401)
    
    if user.role not in ['admin', 'dean']:
        return HttpResponse('Huquq yo\'q', status=403)

    group_id = request.GET.get('group_id') or request.GET.get('student__group')
    if not group_id:
        return HttpResponse('Group ID talab qilinadi', status=400)

    # Get group and results
    from apps.groups.models import Group
    from apps.subjects.models import Subject
    from apps.students.models import Student
    
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return HttpResponse('Guruh topilmadi', status=404)
    
    students = Student.objects.filter(group=group).order_by('full_name')
    subjects = Subject.objects.filter(tests__groups__id=group_id).distinct().order_by('name')
    
    results_list = []
    for index, student in enumerate(students, start=1):
        student_scores = []
        for subject in subjects:
            # Get all test results for this student and subject
            results = TestResult.objects.filter(
                student=student, 
                test__subject=subject
            ).order_by('started_at')
            
            if results.exists():
                scores = [str(r.score) for r in results]
                student_scores.append(', '.join(scores))
            else:
                student_scores.append('-')
        
        results_list.append({
            'number': str(index),
            'full_name': student.full_name,
            'scores': ' | '.join(student_scores),
            'signature': ''
        })

    context = {
        'university_name': "TOSHKENT IJTIMOIY INNOVATSIYA UNIVERSITETI",
        'Guruh': group.name,
        'Fakultet': "Iqtisodiyot va pedagogika",
        'Kursi': f"{group.course}-kurs",
        'Sana': timezone.now().strftime("%d.%m.%Y"),
        'results_table': results_list,
    }

    template_path = os.path.join(settings.BASE_DIR, 'apps/results/templates/results/docx/vedmost_template.docx')
    
    if not os.path.exists(template_path):
        return HttpResponse('Template topilmadi', status=500)

    try:
        doc = DocxTemplate(template_path)
        doc.render(context)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        filename = f"Vedmost_{group.name}_{timezone.now().strftime('%d-%m-%Y')}.docx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        doc.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Xatolik: {str(e)}', status=500)

    
from django.shortcuts import render
from apps.groups.models import Group
from apps.subjects.models import Subject
from apps.students.models import Student
from django.db.models import Sum, Prefetch
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin

from apps.directions.models import Direction

class VedmostView(LoginRequiredMixin, View):
    def get(self, request):
        user = request.user
        if user.role not in ['admin', 'dean', 'teacher']:
             from django.http import HttpResponseForbidden
             return HttpResponseForbidden("Ruxsat yo'q")

        # 1. Base Data
        directions = Direction.objects.all().order_by('name')
        
        # 2. Get Filter Params
        direction_id = request.GET.get('direction_id')
        course = request.GET.get('course')
        group_id = request.GET.get('group_id')
        
        context = {
            'page': 'vedmost',
            'directions': directions,
            'selected_direction_id': int(direction_id) if direction_id else None,
            'selected_course': int(course) if course else None,
            'selected_group_id': int(group_id) if group_id else None,
            'courses': [], # Will populate if direction is selected
            'groups': [],  # Will populate if course is selected
        }
        
        # 3. Dependent Logic
        if direction_id:
            try:
                selected_direction = Direction.objects.get(id=direction_id)
                context['selected_direction'] = selected_direction
                
                # Fetch available courses for this direction (based on existing groups)
                # Or just hardcode 1-4 if needed, but better to check groups
                # Since Group.direction is a name, we filter by name
                available_courses = Group.objects.filter(direction=selected_direction.name).values_list('course', flat=True).distinct().order_by('course')
                context['courses'] = available_courses
                
                if course:
                    # Fetch Groups for this direction + course
                    groups = Group.objects.filter(direction=selected_direction.name, course=course).order_by('name')
                    context['groups'] = groups
                    
                    if group_id:
                        context['selected_group'] = groups.filter(id=group_id).first()
                        
                        if context['selected_group']:
                             # LOGIC TO GENERATE REPORT (Copied/Adapted from original)
                            students = Student.objects.filter(group_id=group_id).order_by('full_name')
                            subjects = Subject.objects.filter(tests__groups__id=group_id).distinct().order_by('name')
                            
                            report = []
                            all_results = TestResult.objects.filter(
                                student__group_id=group_id,
                                test__subject__in=subjects
                            ).select_related('test', 'test__subject', 'student').order_by('started_at')
                            
                            results_map = {} 
                            for res in all_results:
                                s_id = res.student.id
                                sub_id = res.test.subject.id
                                
                                if s_id not in results_map:
                                    results_map[s_id] = {}
                                if sub_id not in results_map[s_id]:
                                    results_map[s_id][sub_id] = []
                                    
                                results_map[s_id][sub_id].append(str(res.score))
                            
                            
                            for student in students:
                                student_row = {'student': student, 'scores': []}
                                s_map = results_map.get(student.id, {})
                                
                                for subject in subjects:
                                    scores_list = s_map.get(subject.id, [])
                                    scores_str = ", ".join(scores_list) if scores_list else ""
                                    student_row['scores'].append({'subject_id': subject.id, 'value': scores_str})
                                
                                report.append(student_row)

                            context['subjects'] = subjects
                            context['report'] = report

            except Direction.DoesNotExist:
                pass


        return render(request, 'vedmost_list_v5.html', context)


class JamlanmaQaytnomaView(View):
    def get(self, request):
        try:
            if not request.user.is_authenticated:
                from django.shortcuts import redirect
                return redirect('/login/')

            user = request.user
            if user.role == 'student':
                from django.http import HttpResponseForbidden
                return HttpResponseForbidden("Ruxsat yo'q")
                
            if user.role != 'admin':
                from apps.accounts.models import ModuleAccess
                if not ModuleAccess.objects.filter(user=user, module='vedmost', can_view=True).exists():
                    from django.http import HttpResponseForbidden
                    return HttpResponseForbidden("Ruxsat yo'q (Module Access Denied)")

            # 1. Base Data
            directions = Direction.objects.all().order_by('name')
            
            # 2. Get Filter Params
            direction_id = request.GET.get('direction_id')
            course = request.GET.get('course')
            group_id = request.GET.get('group_id')
            export_excel = request.GET.get('export_excel')
            
            context = {
                'page': 'jamlanma_qaytnoma',
                'directions': directions,
                'selected_direction_id': int(direction_id) if direction_id else None,
                'selected_course': int(course) if course else None,
                'selected_group_id': int(group_id) if group_id else None,
                'courses': [],
                'groups': [],
            }
            
            # 3. Dependent Logic
            if direction_id:
                try:
                    selected_direction = Direction.objects.get(id=direction_id)
                    context['selected_direction'] = selected_direction
                    
                    # Fetch available courses
                    available_courses = Group.objects.filter(direction=selected_direction.name).values_list('course', flat=True).distinct().order_by('course')
                    context['courses'] = available_courses
                    
                    if course:
                        # Fetch Groups
                        groups = Group.objects.filter(direction=selected_direction.name, course=course).order_by('name')
                        context['groups'] = groups
                        
                        if group_id:
                            try:
                                selected_group = Group.objects.get(id=group_id)
                                context['selected_group'] = selected_group
                                
                                # 4. Generate Report
                                students = Student.objects.filter(group=selected_group).order_by('full_name')
                                
                                subjects = Subject.objects.filter(
                                    tests__results__student__group=selected_group
                                ).distinct().order_by('name')
                                
                                # Pre-fetch results
                                all_results = TestResult.objects.filter(
                                    student__group=selected_group,
                                    test__subject__in=subjects
                                ).select_related('test', 'test__subject', 'student')
                                
                                results_map = {}
                                for res in all_results:
                                    s_id = res.student.id
                                    sub_id = res.test.subject.id
                                    
                                    if s_id not in results_map: results_map[s_id] = {}
                                    if sub_id not in results_map[s_id]: results_map[s_id][sub_id] = []
                                    
                                    results_map[s_id][sub_id].append(str(res.score))

                                report = []
                                for index, student in enumerate(students, 1):
                                    row = {
                                        'number': index,
                                        'student': student,
                                        'scores': []
                                    }
                                    s_map = results_map.get(student.id, {})
                                    
                                    for subject in subjects:
                                        scores_list = s_map.get(subject.id, [])
                                        scores_str = ", ".join(scores_list) if scores_list else ""
                                        row['scores'].append({'subject_id': subject.id, 'value': scores_str})
                                    
                                    report.append(row)
                                    
                                context['subjects'] = subjects
                                context['report'] = report
                                
                                # Calculate Statistics
                                stats = []
                                total_students_count = students.count()
                                
                                for subject in subjects:
                                    subject_stats = {
                                        'subject': subject,
                                        'total': total_students_count,
                                        'participated': 0,
                                        'not_participated': 0,
                                        'grade_5': 0,
                                        'grade_4': 0,
                                        'grade_3': 0,
                                        'failed': 0
                                    }
                                    
                                    # Get all results for this subject/group
                                    subject_results = [r for r in all_results if r.test.subject.id == subject.id]
                                    
                                    unique_student_ids = set()
                                    for res in subject_results:
                                        unique_student_ids.add(res.student.id)
                                        
                                        grade = 0
                                        if res.max_score == 50:
                                            if res.score >= 40:
                                                grade = 5
                                            elif res.score >= 35:
                                                grade = 4
                                            elif res.score >= 30:
                                                grade = 3
                                            else:
                                                grade = 2
                                        else:
                                            p = res.percentage
                                            if p >= 90:
                                                grade = 5
                                            elif p >= 70:
                                                grade = 4
                                            elif p >= 60:
                                                grade = 3
                                            else:
                                                grade = 2

                                        if grade == 5:
                                            subject_stats['grade_5'] += 1
                                        elif grade == 4:
                                            subject_stats['grade_4'] += 1
                                        elif grade == 3:
                                            subject_stats['grade_3'] += 1
                                        else:
                                            subject_stats['failed'] += 1
                                    
                                    subject_stats['participated'] = len(unique_student_ids) 
                                    subject_stats['not_participated'] = total_students_count - subject_stats['participated']
                                    
                                    stats.append(subject_stats)
                                    
                                context['stats'] = stats
                                
                                # 5. Export Logic
                                if export_excel == 'true':
                                    return self.export_to_excel(selected_group, subjects, report, stats)

                            except Group.DoesNotExist:
                                pass

                except Direction.DoesNotExist:
                    pass

            return render(request, 'results/jamlanma_qaytnoma.html', context)

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            return HttpResponse(
                f"<pre>ERROR: {str(e)}\n\n{tb}</pre>",
                status=500,
                content_type='text/html'
            )

    def export_to_excel(self, group, subjects, report, stats):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{group.name} - Jamlanma"
        
        # Styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bg_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Header 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(subjects))
        ws['A1'] = f"Guruh: {group.name} | Kurs: {group.course} | Yo'nalish: {group.direction}"
        ws['A1'].font = bold_font
        ws['A1'].alignment = center_align

        # Header 2: Column Names
        headers = ["№", "F.I.Sh", "ID"] + [s.name for s in subjects]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data Rows
        for row_data in report:
            student = row_data['student']
            scores = [s['value'] for s in row_data['scores']]
            
            row_cells = [row_data['number'], student.full_name, student.student_id] + scores
            ws.append(row_cells)
            
            # Apply border to data cells
            for col_num in range(1, len(row_cells) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = thin_border
                if col_num > 3: # Score columns
                    cell.alignment = center_align
        
        # Statistics Section
        ws.append([]) # Empty row
        start_row = ws.max_row + 1
        
        stat_labels = [
            ("Jami talabalar", 'total'),
            ("Qatnashdi", 'participated'),
            ("Qatnashmadi", 'not_participated'),
            ("5 baho (A'lo)", 'grade_5'),
            ("4 baho (Yaxshi)", 'grade_4'),
            ("3 baho (Qoniqarli)", 'grade_3'),
            ("Yiqildi (Qoniqarsiz)", 'failed'),
        ]

        for label, key in stat_labels:
            row_cells = ["", label, ""] 
            for stat in stats:
                row_cells.append(stat[key])
            
            ws.append(row_cells)
            
            # Formatting
            current_row = ws.max_row
            label_cell = ws.cell(row=current_row, column=2)
            label_cell.font = bold_font
            label_cell.border = thin_border
            
            for i in range(len(stats)):
                val_cell = ws.cell(row=current_row, column=4 + i)
                val_cell.alignment = center_align
                val_cell.border = thin_border


        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        for i in range(len(subjects)):
            col_letter = openpyxl.utils.get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        timestamp = timezone.now().strftime('%d_%m_%Y_%H_%M_%S')
        filename = f"Jamlanma_{group.name}_{timestamp}.xlsx"
        safe_filename = escape_uri_path(filename)
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe_filename}"
        wb.save(response)
        return response

def result_list_view(request):
    return render(request, 'crud_list.html', {'page': 'results'})