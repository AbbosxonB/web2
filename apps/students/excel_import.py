import openpyxl
from django.contrib.auth import get_user_model
from .models import Student
from apps.groups.models import Group

User = get_user_model()

def import_students_from_excel(file):
    # Use read_only=True for memory efficiency, data_only=True to get values not formulas
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active
    # ... rest of code
    
    from apps.directions.models import Direction
    
    count = 0
    errors = []
    
    print(f"Starting import. Sheet max row: {sheet.max_row}")
    # when read_only is True, max_row might be None or approximate, but iter_rows works.
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        print(f"Processing row: {row}")
        if not row or not row[0]: # Skip empty
            print("Skipping empty row")
            continue
            
        full_name = row[0]
        student_id = str(row[1])
        # ... rest of valid vars
        group_name = row[2]
        course = int(row[3]) if row[3] else 1
        direction_name = row[4] or 'Default'
        education_form = row[5].lower() if row[5] else 'kunduzgi'
        phone = str(row[6]) if row[6] else ''
        email = str(row[7]) if len(row) > 7 and row[7] else ''
        username = str(row[8]) if len(row) > 8 and row[8] else student_id
        password = str(row[9]) if len(row) > 9 and row[9] else student_id

        if not student_id or Student.objects.filter(student_id=student_id).exists():
            print(f"Skipping existing student or missing ID: {student_id}")
            continue
        
        # Sync Direction
        # Sync Direction
        direction_obj = None
        if direction_name:
            direction_obj = Direction.objects.filter(name__iexact=direction_name).first()
            
            if not direction_obj:
                # Create new with unique code
                import uuid
                # Ensure code is unique and within limit. 
                code_base = direction_name[:20].upper().replace(' ', '_')
                unique_suffix = str(uuid.uuid4())[:8]
                code = f"{code_base}_{course}_{unique_suffix}"
                direction_obj = Direction.objects.create(name=direction_name, code=code[:50])

        # Sync Group
        group = None
        if group_name:
            group, created = Group.objects.get_or_create(
                name=group_name,
                defaults={
                    'course': course,
                    'direction': direction_name,
                    'education_form': education_form
                }
            )

        # Create User
        # Find or Create User
        user = User.objects.filter(username=username).first()
        if not user:
            user = User.objects.create_user(username=username, password=password, role='student')
        else:
             print(f"User {username} exists. Using existing user.")
             # Update password if needed? For now, keep existing.
             pass
            
        Student.objects.create(
            user=user,
            student_id=student_id,
            full_name=full_name,
            phone=phone,
            email=email,
            group=group,
            course=course,
            direction=direction_name,
            education_form=education_form
        )
        count += 1
        
    return count
