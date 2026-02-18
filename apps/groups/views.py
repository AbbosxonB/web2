from rest_framework import viewsets, permissions, filters, pagination
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Group
from .serializers import GroupSerializer

from django.shortcuts import render

class GroupPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    from apps.accounts.granular_permissions import GranularPermission
    permission_classes = [permissions.IsAuthenticated, GranularPermission]
    module_name = 'groups'
    pagination_class = GroupPagination
    filter_backends = [filters.SearchFilter]
    ordering = ['name']

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log_action('create', instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        # Detect changes roughly
        changed_fields = [k for k in self.request.data.keys() if k not in ['id', 'csrfmiddlewaretoken']]
        extra = f"O'zgarganlar: {', '.join(changed_fields)}" if changed_fields else "Tahrirlandi"
        self._log_action('update', instance, extra_details=extra)

    def perform_destroy(self, instance):
        self._log_action('delete', instance)
        instance.delete()

    def _log_action(self, action_type, instance, extra_details=None):
        from apps.logs.models import SystemLog
        
        action_map = {
            'create': "Guruh yaratildi",
            'update': "Guruh tahrirlandi",
            'delete': "Guruh o'chirildi"
        }
        
        details = f"Guruh: {instance.name} (ID: {instance.id})"
        if action_type == 'delete':
            details = f"Guruh: {instance.name} (ID: {instance.id})"
            
        if extra_details:
            details += f". {extra_details}"
            
        ip = self.request.META.get('REMOTE_ADDR')
        
        SystemLog.objects.create(
            user=self.request.user,
            action=action_map.get(action_type, action_type),
            details=details,
            ip_address=ip
        )
    search_fields = ['name', 'direction']

    def get_queryset(self):
        queryset = Group.objects.all()
        course = self.request.query_params.get('course')
        direction = self.request.query_params.get('direction')
        
        is_active = self.request.query_params.get('is_system_active')
        
        if course:
            queryset = queryset.filter(course=course)
        if direction:
            queryset = queryset.filter(direction__icontains=direction)
        if is_active is not None:
             # Convert string 'true'/'false' to boolean
             is_active_bool = is_active.lower() == 'true'
             queryset = queryset.filter(is_system_active=is_active_bool)
            
        return queryset

    @action(detail=False, methods=['post'])
    def bulk_action(self, request):
        user = request.user
        if user.role != 'admin' and (not user.permissions or not any(p.module == 'groups' and p.can_update for p in user.permissions)):
             return Response({'error': 'Huquq yo\'q'}, status=403)
        
        action = request.data.get('action')
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({'error': 'Hech narsa tanlanmadi'}, status=400)
            
        queryset = Group.objects.filter(id__in=ids)
        
        if action == 'delete':
             if user.role != 'admin' and (not user.permissions or not any(p.module == 'groups' and p.can_delete for p in user.permissions)):
                 return Response({'error': 'O\'chirish uchun huquq yo\'q'}, status=403)
             queryset.delete()
             return Response({'status': 'deleted', 'count': len(ids)})
             
        elif action == 'activate':
            queryset.update(is_system_active=True)
            return Response({'status': 'activated', 'count': len(ids)})
            
        elif action == 'deactivate':
            queryset.update(is_system_active=False)
            return Response({'status': 'deactivated', 'count': len(ids)})
            
        return Response({'error': 'Noto\'g\'ri amal'}, status=400)

    @action(detail=False, methods=['post'])
    def import_data(self, request):
        import openpyxl
        from apps.directions.models import Direction
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Fayl tanlanmadi'}, status=400)
            
        try:
            wb = openpyxl.load_workbook(file_obj)
            sheet = wb.active
            
            count = 0
            
            # Format: 
            # Nomi | Kurs | Yo'nalish | Ta'lim Shakli
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]: continue
                
                name = str(row[0]).strip()
                course = int(row[1]) if row[1] else 1
                dir_val = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                edu_form = str(row[3]).strip() if len(row) > 3 and row[3] else 'kunduzgi'
                
                if Group.objects.filter(name=name).exists():
                    continue 

                # Logic: Find or Create Direction
                direction_obj = None
                if dir_val:
                    # Try by Name first
                    d = Direction.objects.filter(name__iexact=dir_val).first()
                    if not d:
                        # Try by Code
                        d = Direction.objects.filter(code__iexact=dir_val).first()
                    
                    if not d:
                        # Not found -> Create New
                        # Generate simple code if not provided
                        new_code = dir_val.upper()[:10]
                        # Ensure code unique
                        if Direction.objects.filter(code=new_code).exists():
                            new_code = f"{new_code}_{count}"
                        
                        d = Direction.objects.create(name=dir_val, code=new_code)
                        # Log direction creation? Optional but good practice
                        
                    direction_obj = d
                
                # Create Group
                # Model uses CharField for direction name usually in this legacy codebase or joined?
                # Looking at models.py earlier: direction = models.CharField(max_length=255)
                # It does NOT use ForeignKey! 
                # Wait, wait. Step 44 showed: `direction = models.CharField(max_length=255)`
                # So we just save the STRING name.
                # BUT, the user prompt implies "Direction" concept exists.
                # If the model is just CharField, then my lookup logic for Direction model is theoretically correct but practically maybe redundant if we just store string?
                # However, the user said "yo'nalishni olsin... kodini uzi avtomatik quyib bersin".
                # If the group `direction` field is just a string name, then `d.name` is what we put there.
                # AND `d.name` comes from the Excel `dir_val`.
                # So if we create a Direction object, it's for consistency in the "Directions" table, 
                # even if Group just stores the name string.
                # Let's verify if Group links to Direction. 
                # Step 44: `direction = models.CharField(max_length=255)`
                # BUT Step 45: `class Direction(models.Model)...`
                # So they are loosely coupled by name.
                
                # The User requirement: "topilmasa usha excel dagi orqali yaratish kerak" (if not found, create via excel one)
                # So yes, I MUST create the Direction object if it doesn't exist, 
                # and then save its name into the Group.direction field.
                
                final_dir_name = direction_obj.name if direction_obj else dir_val
                
                instance = Group.objects.create(
                    name=name,
                    course=course,
                    direction=final_dir_name, # Storing name as per model definition
                    education_form=edu_form
                )
                self._log_action('create', instance, 'Import orqali yaratildi')
                count += 1
                
            return Response({'status': 'success', 'count': count})
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def sample_file(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Namuna"
        
        # Headers
        ws.append(['Nomi', 'Kurs', "Yo'nalish", "Ta'lim Shakli"])
        
        # Sample Data
        ws.append(['911-21', 4, 'Dasturiy Injiniring', 'kunduzgi'])
        ws.append(['912-21', 4, 'Axborot Xavfsizligi', 'sirtqi'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=guruhlar_namuna.xlsx'
        
        wb.save(response)
        return response

def group_list_view(request):
    return render(request, 'crud_list.html', {'page': 'groups'})
